import os
import sys
import glob
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from PIL import Image, ImageDraw, ImageFont, ImageTk

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Theme Colors (Premium "One Dark" Slate Theme)
THEME = {
    "bg": "#1e222b",            # Main window background
    "card_bg": "#282c34",       # Panel/card background
    "text": "#abb2bf",          # Subtext/regular labels (slate gray)
    "text_light": "#ffffff",    # Main headings/bright text
    "accent": "#61afef",        # Blue accent
    "accent_hover": "#4fa3e3",  # Blue hover
    "success": "#98c379",       # Green accent (run button)
    "success_hover": "#8ab66c", # Green hover
    "danger": "#e06c75",        # Red for remove buttons
    "danger_hover": "#c75f6a",  # Red hover
    "border": "#3e4452",        # Border lines
    "entry_bg": "#21252b",      # Inputs/lists background
    "field_bg": "#2c313a",      # Extra field row background
}

FONTS = {
    "Arial Bold": "arialbd.ttf",
    "Arial Regular": "arial.ttf",
    "Segoe UI Bold": "segoeuib.ttf",
    "Segoe UI Regular": "segoeui.ttf",
    "Calibri Bold": "calibrib.ttf",
    "Calibri Regular": "calibri.ttf",
    "Times New Roman Bold": "timesbd.ttf",
    "Courier New Bold": "courbd.ttf",
}

# Columns in Full Price List sheet that are useful for labeling
XLSX_SHEET_NAME = "Full Price List"
XLSX_ART_COL = "Art no."

def hex_to_rgb(hex_str):
    hex_str = hex_str.lstrip('#')
    return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))

def rgb_to_hex(rgb):
    return '#{:02x}{:02x}{:02x}'.format(*rgb)


class LabelerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Shawl Image Art-Number Labeler")
        self.root.geometry("1150x800")
        self.root.configure(bg=THEME["bg"])

        # State variables
        self.input_dir = tk.StringVar(value=os.path.abspath("./input_shawls") if os.path.exists("./input_shawls") else "")
        self.output_dir = tk.StringVar(value=os.path.abspath("./output_shawls") if os.path.exists("./output_shawls") else "")

        # Default settings: White bg, Black text, size 16
        self.font_size_var = tk.IntVar(value=16)
        self.text_color_var = tk.StringVar(value="#000000")  # Black
        self.bg_color_var = tk.StringVar(value="#ffffff")    # White
        self.bg_opacity_var = tk.DoubleVar(value=1.0)        # Opaque
        self.position_var = tk.StringVar(value="Top-Left")
        self.font_name_var = tk.StringVar(value="Arial Bold")
        self.template_var = tk.StringVar(value="{filename}")
        self.padding_var = tk.IntVar(value=10)
        self.margin_var = tk.IntVar(value=20)

        # XLSX / Price data
        self.xlsx_path_var = tk.StringVar(value="")
        self.xlsx_data = {}          # art_no (str) -> {col_name: value, ...}
        self.xlsx_columns = []       # list of column header strings from xlsx
        self.xlsx_sheets = []        # list of sheet names found in the workbook
        self.sheet_name_var = tk.StringVar(value="")  # currently selected sheet
        self.sheet_menu = None       # OptionMenu widget — set during UI build
        self.extra_fields = []       # list of {"frame": Frame, "text_var": StringVar, "col_var": StringVar}
        self.extra_fields_frame = None  # container widget — set during UI build

        self.image_files = []
        self.current_preview_index = 0
        self.preview_image_raw = None
        self.preview_debounce_id = None
        self.processing = False

        # Create UI Layout
        self.create_layout()

        # Initial scan if input directory is preset
        if self.input_dir.get():
            self.scan_input_folder(silent=True)

    # ─────────────────────────────────────────────────────────────────────────
    # Layout
    # ─────────────────────────────────────────────────────────────────────────
    def create_layout(self):
        # 1. Header Banner
        header = tk.Frame(self.root, bg=THEME["card_bg"], height=60, bd=0)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        title_label = tk.Label(
            header,
            text="SHAWL IMAGE ART-NUMBER LABELER",
            fg=THEME["text_light"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 14, "bold")
        )
        title_label.pack(side="left", padx=20, pady=15)

        subtitle_label = tk.Label(
            header,
            text="Automated Client-Photo Branding Utility  •  with XLSX Price Data",
            fg=THEME["text"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 10, "italic")
        )
        subtitle_label.pack(side="left", padx=10, pady=18)

        # 2. Main Work Area (Panels side by side)
        main_pane = tk.Frame(self.root, bg=THEME["bg"])
        main_pane.pack(fill="both", expand=True, padx=15, pady=15)

        # Left Panel (Settings) — Width 460
        self.left_panel = tk.Frame(main_pane, bg=THEME["card_bg"], width=460, bd=1, relief="solid", highlightbackground=THEME["border"])
        self.left_panel.pack(side="left", fill="both", padx=(0, 10))
        self.left_panel.pack_propagate(False)

        # Right Panel (Preview) — Fills remaining space
        self.right_panel = tk.Frame(main_pane, bg=THEME["card_bg"], bd=1, relief="solid", highlightbackground=THEME["border"])
        self.right_panel.pack(side="right", fill="both", expand=True)

        self.build_settings_panel()
        self.build_preview_panel()

    # ─────────────────────────────────────────────────────────────────────────
    # Settings Panel
    # ─────────────────────────────────────────────────────────────────────────
    def build_settings_panel(self):
        title = tk.Label(
            self.left_panel,
            text="CONFIGURATION SETTINGS",
            fg=THEME["text_light"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 11, "bold")
        )
        title.pack(anchor="w", padx=15, pady=(15, 10))

        divider = tk.Frame(self.left_panel, height=1, bg=THEME["border"])
        divider.pack(fill="x", padx=15, pady=(0, 15))

        # Scrollable container
        scroll_container = tk.Canvas(self.left_panel, bg=THEME["card_bg"], bd=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.left_panel, orient="vertical", command=scroll_container.yview)
        self.settings_content = tk.Frame(scroll_container, bg=THEME["card_bg"])

        self.settings_content.bind(
            "<Configure>",
            lambda e: scroll_container.configure(scrollregion=scroll_container.bbox("all"))
        )

        window_id = scroll_container.create_window((0, 0), window=self.settings_content, anchor="nw")
        scroll_container.bind("<Configure>", lambda e: scroll_container.itemconfig(window_id, width=e.width))

        scroll_container.configure(yscrollcommand=scrollbar.set)

        scroll_container.pack(side="left", fill="both", expand=True, padx=(15, 0), pady=(0, 10))
        scrollbar.pack(side="right", fill="y", padx=(0, 5), pady=(0, 10))

        # ── Group 1: Directories ─────────────────────────────────────────────
        self.create_group_label(self.settings_content, "Folders Setup")

        tk.Label(self.settings_content, text="Input Folder (containing shawls):", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 2))
        input_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        input_f.pack(fill="x", pady=(0, 10))
        self.input_entry = self.create_entry(input_f, textvariable=self.input_dir, width=32)
        self.input_entry.pack(side="left", fill="x", expand=True)
        self.input_entry.bind("<KeyRelease>", lambda e: self.scan_input_folder())
        self.create_button(input_f, "Browse", self.browse_input_dir, THEME["border"]).pack(side="right", padx=(5, 0))

        tk.Label(self.settings_content, text="Output Folder (labeled shawls):", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 2))
        output_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        output_f.pack(fill="x", pady=(0, 15))
        self.output_entry = self.create_entry(output_f, textvariable=self.output_dir, width=32)
        self.output_entry.pack(side="left", fill="x", expand=True)
        self.create_button(output_f, "Browse", self.browse_output_dir, THEME["border"]).pack(side="right", padx=(5, 0))

        # ── Group 2: Label Styling ────────────────────────────────────────────
        self.create_group_label(self.settings_content, "Label Style Customization")

        tk.Label(self.settings_content, text="Art-Number Text Template (use {filename} for Art Number):", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 2))
        self.template_entry = self.create_entry(self.settings_content, textvariable=self.template_var, width=40)
        self.template_entry.pack(fill="x", pady=(0, 10))
        self.template_entry.bind("<KeyRelease>", self.schedule_preview_update)

        # Font Dropdown
        tk.Label(self.settings_content, text="Font Style:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 2))
        self.font_menu = self.create_option_menu(self.settings_content, self.font_name_var, list(FONTS.keys()), self.schedule_preview_update)
        self.font_menu.pack(fill="x", pady=(0, 10))

        # Font Size
        size_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        size_f.pack(fill="x", pady=(5, 10))
        tk.Label(size_f, text="Font Size:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(side="left")
        self.size_spin = tk.Spinbox(size_f, from_=8, to=300, textvariable=self.font_size_var, width=5, bg=THEME["entry_bg"], fg=THEME["text_light"], bd=1, relief="solid", highlightthickness=0, font=("Segoe UI", 10), command=self.schedule_preview_update)
        self.size_spin.pack(side="right", padx=(5, 0))
        self.size_spin.bind("<KeyRelease>", self.schedule_preview_update)
        self.size_slider = self.create_scale(size_f, 8, 200, self.font_size_var)
        self.size_slider.pack(side="right", fill="x", expand=True, padx=(10, 5))

        # Colors Row
        color_row = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        color_row.pack(fill="x", pady=(5, 10))

        # Font Color
        font_c_f = tk.Frame(color_row, bg=THEME["card_bg"])
        font_c_f.pack(side="left", fill="x", expand=True)
        tk.Label(font_c_f, text="Font Color:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w")
        font_picker_f = tk.Frame(font_c_f, bg=THEME["card_bg"])
        font_picker_f.pack(fill="x", pady=2)
        self.font_color_preview = tk.Canvas(font_picker_f, width=25, height=20, bg=self.text_color_var.get(), bd=1, relief="solid", highlightthickness=0)
        self.font_color_preview.pack(side="left")
        self.create_button(font_picker_f, "Choose", self.choose_font_color, THEME["border"]).pack(side="left", padx=5)

        # Background Color
        bg_c_f = tk.Frame(color_row, bg=THEME["card_bg"])
        bg_c_f.pack(side="right", fill="x", expand=True)
        tk.Label(bg_c_f, text="Box Background:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w")
        bg_picker_f = tk.Frame(bg_c_f, bg=THEME["card_bg"])
        bg_picker_f.pack(fill="x", pady=2)
        self.bg_color_preview = tk.Canvas(bg_picker_f, width=25, height=20, bg=self.bg_color_var.get(), bd=1, relief="solid", highlightthickness=0)
        self.bg_color_preview.pack(side="left")
        self.create_button(bg_picker_f, "Choose", self.choose_bg_color, THEME["border"]).pack(side="left", padx=5)

        # Background Opacity
        opacity_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        opacity_f.pack(fill="x", pady=(5, 10))
        tk.Label(opacity_f, text="Background Opacity:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(side="left")
        self.opacity_slider = self.create_scale(opacity_f, 0.0, 1.0, self.bg_opacity_var)
        self.opacity_slider.configure(resolution=0.1)
        self.opacity_slider.pack(side="right", fill="x", expand=True, padx=(10, 0))
        self.opacity_slider.bind("<ButtonRelease-1>", self.schedule_preview_update)

        # Padding & Margin
        metrics_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        metrics_f.pack(fill="x", pady=5)

        pad_f = tk.Frame(metrics_f, bg=THEME["card_bg"])
        pad_f.pack(side="left", fill="x", expand=True, padx=(0, 5))
        tk.Label(pad_f, text="Box Padding:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w")
        self.pad_slider = self.create_scale(pad_f, 0, 50, self.padding_var)
        self.pad_slider.pack(fill="x")
        self.pad_slider.bind("<ButtonRelease-1>", self.schedule_preview_update)

        margin_f = tk.Frame(metrics_f, bg=THEME["card_bg"])
        margin_f.pack(side="right", fill="x", expand=True, padx=(5, 0))
        tk.Label(margin_f, text="Edge Margin:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w")
        self.margin_slider = self.create_scale(margin_f, 0, 150, self.margin_var)
        self.margin_slider.pack(fill="x")
        self.margin_slider.bind("<ButtonRelease-1>", self.schedule_preview_update)

        # Position Dropdown
        tk.Label(self.settings_content, text="Text Position:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(5, 2))
        self.pos_menu = self.create_option_menu(self.settings_content, self.position_var, ["Top-Left", "Top-Right", "Bottom-Left", "Bottom-Right", "Center", "Top-Center", "Bottom-Center"], self.schedule_preview_update)
        self.pos_menu.pack(fill="x", pady=(0, 15))

        # ── Group 3: Price Data (XLSX) ────────────────────────────────────────
        self.create_group_label(self.settings_content, "Price Data (XLSX)")

        info_text = "Load the price xlsx to add price/item fields below." if OPENPYXL_AVAILABLE else "⚠ openpyxl not installed. Run: pip install openpyxl"
        tk.Label(self.settings_content, text=info_text, fg=THEME["accent"] if OPENPYXL_AVAILABLE else THEME["danger"], bg=THEME["card_bg"], font=("Segoe UI", 8, "italic"), wraplength=380, justify="left").pack(anchor="w", pady=(0, 8))

        tk.Label(self.settings_content, text="Price List XLSX File:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(anchor="w", pady=(2, 2))
        xlsx_f = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        xlsx_f.pack(fill="x", pady=(0, 6))
        self.xlsx_entry = self.create_entry(xlsx_f, textvariable=self.xlsx_path_var, width=28)
        self.xlsx_entry.pack(side="left", fill="x", expand=True)
        self.create_button(xlsx_f, "Browse", self.browse_xlsx_file, THEME["border"]).pack(side="right", padx=(5, 0))

        load_row = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        load_row.pack(fill="x", pady=(0, 6))
        self.create_button(load_row, "⟳  Peek Sheets", self.peek_xlsx_sheets, THEME["accent"], font=("Segoe UI", 9, "bold")).pack(side="left")
        self.xlsx_status_label = tk.Label(load_row, text="No file loaded.", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 8, "italic"))
        self.xlsx_status_label.pack(side="left", padx=(10, 0))

        # Sheet selector
        sheet_row = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        sheet_row.pack(fill="x", pady=(0, 4))
        tk.Label(sheet_row, text="Sheet:", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 9)).pack(side="left")
        self.sheet_name_var.set("— load file first —")
        self.sheet_menu = tk.OptionMenu(sheet_row, self.sheet_name_var, "— load file first —")
        self.sheet_menu.config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["border"], activeforeground=THEME["text_light"],
            bd=1, relief="solid", highlightthickness=0, font=("Segoe UI", 9)
        )
        self.sheet_menu["menu"].config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["accent"], activeforeground=THEME["text_light"],
            font=("Segoe UI", 9)
        )
        self.sheet_menu.pack(side="left", padx=(8, 0), fill="x", expand=True)

        load_data_row = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        load_data_row.pack(fill="x", pady=(0, 8))
        self.create_button(load_data_row, "✓  Load Sheet Data", self.load_xlsx, THEME["success"], font=("Segoe UI", 9, "bold")).pack(side="left")

        # ── Group 4: Additional Label Fields ─────────────────────────────────
        self.create_group_label(self.settings_content, "Additional Label Fields")

        hint = tk.Label(
            self.settings_content,
            text="Each field adds a new line to the label.\nFormat: [Prefix Text]  [XLSX Column Value]",
            fg=THEME["text"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 8, "italic"),
            justify="left"
        )
        hint.pack(anchor="w", pady=(0, 8))

        hdr = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        hdr.pack(fill="x", pady=(0, 4))
        tk.Label(hdr, text="Prefix Text", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 8, "bold"), width=14, anchor="w").pack(side="left")
        tk.Label(hdr, text="XLSX Column", fg=THEME["text"], bg=THEME["card_bg"], font=("Segoe UI", 8, "bold"), anchor="w").pack(side="left", padx=(4, 0))

        self.extra_fields_frame = tk.Frame(self.settings_content, bg=THEME["card_bg"])
        self.extra_fields_frame.pack(fill="x", pady=(0, 8))

        self.create_button(
            self.settings_content,
            "+ Add Field",
            self.add_field_row,
            THEME["accent"],
            font=("Segoe UI", 9, "bold")
        ).pack(anchor="w", pady=(0, 15))

        # ── Group 5: Process Output ───────────────────────────────────────────
        self.create_group_label(self.settings_content, "Process Output")

        self.btn_process = self.create_button(
            self.settings_content,
            "PROCESS ALL SHAWL IMAGES",
            self.start_batch_processing,
            THEME["success"],
            font=("Segoe UI", 11, "bold")
        )
        self.btn_process.pack(fill="x", pady=(10, 15))

        self.progress = ttk.Progressbar(self.settings_content, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", pady=(0, 5))

        self.status_label = tk.Label(
            self.settings_content,
            text="Ready. Select folders to begin.",
            fg=THEME["text"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 9)
        )
        self.status_label.pack(anchor="w", pady=(0, 15))

        # MouseWheel scrolling
        def _on_mousewheel(event):
            scroll_container.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def bind_mouse_wheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            widget.bind("<Button-4>", lambda e: scroll_container.yview_scroll(-1, "units"))
            widget.bind("<Button-5>", lambda e: scroll_container.yview_scroll(1, "units"))
            for child in widget.winfo_children():
                bind_mouse_wheel(child)

        bind_mouse_wheel(scroll_container)
        bind_mouse_wheel(self.settings_content)

    # ─────────────────────────────────────────────────────────────────────────
    # Preview Panel
    # ─────────────────────────────────────────────────────────────────────────
    def build_preview_panel(self):
        title_f = tk.Frame(self.right_panel, bg=THEME["card_bg"])
        title_f.pack(fill="x", padx=15, pady=(15, 10))

        tk.Label(title_f, text="LIVE IMAGE PREVIEW", fg=THEME["text_light"], bg=THEME["card_bg"], font=("Segoe UI", 11, "bold")).pack(side="left")

        self.image_info_label = tk.Label(title_f, text="", fg=THEME["accent"], bg=THEME["card_bg"], font=("Segoe UI", 9, "bold"))
        self.image_info_label.pack(side="right")

        divider = tk.Frame(self.right_panel, height=1, bg=THEME["border"])
        divider.pack(fill="x", padx=15, pady=(0, 15))

        self.preview_canvas_frame = tk.Frame(self.right_panel, bg=THEME["entry_bg"], bd=1, relief="solid", highlightbackground=THEME["border"])
        self.preview_canvas_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        self.preview_label = tk.Label(
            self.preview_canvas_frame,
            text="Select an Input Folder containing images\nto display a dynamic preview here.",
            fg=THEME["text"],
            bg=THEME["entry_bg"],
            font=("Segoe UI", 11, "italic")
        )
        self.preview_label.pack(fill="both", expand=True)

        self.nav_frame = tk.Frame(self.right_panel, bg=THEME["card_bg"])
        self.nav_frame.pack(fill="x", side="bottom", pady=15, padx=15)

        self.btn_prev = self.create_button(self.nav_frame, "◀ Previous", self.show_prev_image, THEME["border"])
        self.btn_prev.pack(side="left")

        self.preview_counter_label = tk.Label(
            self.nav_frame,
            text="No Images Loaded",
            fg=THEME["text"],
            bg=THEME["card_bg"],
            font=("Segoe UI", 10, "bold")
        )
        self.preview_counter_label.pack(side="left", fill="x", expand=True)

        self.btn_next = self.create_button(self.nav_frame, "Next ▶", self.show_next_image, THEME["border"])
        self.btn_next.pack(side="right")

        self.update_nav_buttons()

    # ─────────────────────────────────────────────────────────────────────────
    # Custom Widget Helpers
    # ─────────────────────────────────────────────────────────────────────────
    def create_group_label(self, parent, text):
        lbl = tk.Label(parent, text=text.upper(), fg=THEME["accent"], bg=THEME["card_bg"], font=("Segoe UI", 9, "bold"))
        lbl.pack(anchor="w", pady=(15, 5))
        line = tk.Frame(parent, height=1, bg=THEME["border"])
        line.pack(fill="x", pady=(0, 8))

    def create_entry(self, parent, textvariable=None, width=20):
        entry = tk.Entry(
            parent,
            textvariable=textvariable,
            bg=THEME["entry_bg"],
            fg=THEME["text_light"],
            insertbackground=THEME["text_light"],
            bd=1,
            relief="solid",
            highlightbackground=THEME["border"],
            highlightcolor=THEME["accent"],
            highlightthickness=1,
            font=("Segoe UI", 10),
            width=width
        )
        return entry

    def create_button(self, parent, text, command, bg_color, fg_color=THEME["text_light"], font=("Segoe UI", 9, "bold")):
        btn = tk.Button(
            parent, text=text, command=command,
            bg=bg_color, fg=fg_color,
            font=font,
            activebackground=bg_color, activeforeground=fg_color,
            bd=0, relief="flat", padx=12, pady=6, cursor="hand2"
        )
        def on_enter(e):
            if bg_color == THEME["accent"]:
                btn.config(bg=THEME["accent_hover"])
            elif bg_color == THEME["success"]:
                btn.config(bg=THEME["success_hover"])
            elif bg_color == THEME["danger"]:
                btn.config(bg=THEME["danger_hover"])
            else:
                btn.config(bg=THEME["border"])
        def on_leave(e):
            btn.config(bg=bg_color)
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def create_scale(self, parent, from_, to, variable=None):
        scale = tk.Scale(
            parent, from_=from_, to=to, variable=variable,
            orient="horizontal",
            bg=THEME["card_bg"], fg=THEME["text"],
            troughcolor=THEME["entry_bg"], activebackground=THEME["accent"],
            highlightthickness=0, bd=0,
            font=("Segoe UI", 8), showvalue=True
        )
        scale.bind("<ButtonRelease-1>", self.schedule_preview_update)
        return scale

    def create_option_menu(self, parent, variable, values, command=None):
        def menu_selected(*args):
            if command:
                command()
        variable.trace_add("write", menu_selected)
        menu = tk.OptionMenu(parent, variable, *values)
        menu.config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["border"], activeforeground=THEME["text_light"],
            bd=1, relief="solid", highlightthickness=0, font=("Segoe UI", 9)
        )
        menu["menu"].config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["accent"], activeforeground=THEME["text_light"],
            font=("Segoe UI", 9)
        )
        return menu

    # ─────────────────────────────────────────────────────────────────────────
    # XLSX Loading
    # ─────────────────────────────────────────────────────────────────────────
    def browse_xlsx_file(self):
        d = filedialog.askopenfilename(initialdir=self.input_dir.get() or ".", title="Select Price List XLSX", filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if d:
            self.xlsx_path_var.set(os.path.abspath(d))
            self.peek_xlsx_sheets()

    def peek_xlsx_sheets(self):
        """Open the workbook, read sheet names, populate the sheet dropdown."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "openpyxl is not installed.\nRun: pip install openpyxl")
            return
        path = self.xlsx_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Please select a valid XLSX file first.")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            self.xlsx_sheets = wb.sheetnames
            wb.close()
            if not self.xlsx_sheets:
                messagebox.showerror("Error", "No sheets found in the workbook.")
                return
            # Rebuild the sheet dropdown
            menu = self.sheet_menu["menu"]
            menu.delete(0, "end")
            for s in self.xlsx_sheets:
                menu.add_command(label=s, command=lambda v=s: self.sheet_name_var.set(v))
            # Pre-select first sheet (or keep current if still valid)
            if self.sheet_name_var.get() not in self.xlsx_sheets:
                self.sheet_name_var.set(self.xlsx_sheets[0])
            self.xlsx_status_label.config(
                text=f"Found {len(self.xlsx_sheets)} sheet(s). Select one and click 'Load Sheet Data'.",
                fg=THEME["accent"]
            )
        except Exception as e:
            messagebox.showerror("XLSX Error", f"Failed to read workbook:\n{str(e)}")
            self.xlsx_status_label.config(text="✗ Load failed.", fg=THEME["danger"])

    def load_xlsx(self):
        """Read data from the currently selected sheet."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Missing Dependency", "openpyxl is not installed.\nRun: pip install openpyxl")
            return
        path = self.xlsx_path_var.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Error", "Please select a valid XLSX file first.")
            return
        sheet_name = self.sheet_name_var.get()
        if not sheet_name or sheet_name == "— load file first —":
            messagebox.showerror("Error", "Please click 'Peek Sheets' first, then select a sheet.")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                messagebox.showerror("Sheet Not Found", f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                return
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
            self.xlsx_columns = [h for h in headers if h]
            try:
                art_col_idx = headers.index(XLSX_ART_COL)
            except ValueError:
                messagebox.showerror("Column Not Found", f"'{XLSX_ART_COL}' column not found in sheet '{sheet_name}'.\nColumns available: {', '.join(self.xlsx_columns)}")
                return
            self.xlsx_data = {}
            for row in rows_iter:
                if row[art_col_idx] is None:
                    continue
                art_no = str(row[art_col_idx]).strip()
                if not art_no:
                    continue
                row_dict = {}
                for i, col_name in enumerate(headers):
                    if col_name and i < len(row):
                        row_dict[col_name] = row[i]
                self.xlsx_data[art_no] = row_dict
            wb.close()
            count = len(self.xlsx_data)
            self.xlsx_status_label.config(
                text=f"✓ Sheet '{sheet_name}' — {count} items · {len(self.xlsx_columns)} columns",
                fg=THEME["success"]
            )
            self._refresh_field_dropdowns()
            self.schedule_preview_update()
        except Exception as e:
            messagebox.showerror("XLSX Load Error", f"Failed to load sheet:\n{str(e)}")
            self.xlsx_status_label.config(text="✗ Load failed.", fg=THEME["danger"])

    def _refresh_field_dropdowns(self):
        cols = self.xlsx_columns if self.xlsx_columns else ["— load xlsx first —"]
        for field in self.extra_fields:
            menu = field["col_menu"]
            menu["menu"].delete(0, "end")
            for col in cols:
                menu["menu"].add_command(label=col, command=lambda v=col, var=field["col_var"]: var.set(v))
            if field["col_var"].get() not in cols:
                field["col_var"].set(cols[0])

    # ─────────────────────────────────────────────────────────────────────────
    # Dynamic Extra Field Rows
    # ─────────────────────────────────────────────────────────────────────────
    def add_field_row(self):
        cols = self.xlsx_columns if self.xlsx_columns else ["— load xlsx first —"]
        row_frame = tk.Frame(self.extra_fields_frame, bg=THEME["field_bg"], bd=1, relief="solid", highlightbackground=THEME["border"], highlightthickness=1)
        row_frame.pack(fill="x", pady=(0, 6))
        text_var = tk.StringVar(value="")
        prefix_entry = self.create_entry(row_frame, textvariable=text_var, width=12)
        prefix_entry.pack(side="left", padx=(6, 4), pady=5)
        prefix_entry.bind("<KeyRelease>", self.schedule_preview_update)
        col_var = tk.StringVar(value=cols[0])
        col_var.trace_add("write", lambda *a: self.schedule_preview_update())
        col_menu = tk.OptionMenu(row_frame, col_var, *cols)
        col_menu.config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["border"], activeforeground=THEME["text_light"],
            bd=1, relief="solid", highlightthickness=0, font=("Segoe UI", 9), width=14
        )
        col_menu["menu"].config(
            bg=THEME["entry_bg"], fg=THEME["text_light"],
            activebackground=THEME["accent"], activeforeground=THEME["text_light"],
            font=("Segoe UI", 9)
        )
        col_menu.pack(side="left", padx=(0, 4), pady=5, fill="x", expand=True)
        field_entry = {"frame": row_frame, "text_var": text_var, "col_var": col_var, "col_menu": col_menu}
        def make_remover(fe):
            return lambda: self.remove_field_row(fe)
        remove_btn = self.create_button(row_frame, "✕", make_remover(field_entry), THEME["danger"], font=("Segoe UI", 9, "bold"))
        remove_btn.pack(side="right", padx=(0, 6), pady=5)
        self.extra_fields.append(field_entry)
        self.schedule_preview_update()

    def remove_field_row(self, field_entry):
        if field_entry in self.extra_fields:
            self.extra_fields.remove(field_entry)
        field_entry["frame"].destroy()
        self.schedule_preview_update()

    # ─────────────────────────────────────────────────────────────────────────
    # Folder Handling
    # ─────────────────────────────────────────────────────────────────────────
    def browse_input_dir(self):
        d = filedialog.askdirectory(initialdir=self.input_dir.get() or ".")
        if d:
            self.input_dir.set(os.path.abspath(d))
            self.scan_input_folder()

    def browse_output_dir(self):
        d = filedialog.askdirectory(initialdir=self.output_dir.get() or ".")
        if d:
            self.output_dir.set(os.path.abspath(d))

    def scan_input_folder(self, silent=False):
        dir_path = self.input_dir.get()
        if not dir_path or not os.path.exists(dir_path):
            self.image_files = []
            self.current_preview_index = 0
            self.preview_image_raw = None
            self.preview_label.config(text="Select an Input Folder containing images\nto display a dynamic preview here.")
            self.preview_counter_label.config(text="No Images Loaded")
            self.image_info_label.config(text="")
            self.update_nav_buttons()
            return
        valid_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp')
        try:
            files = [f for f in os.listdir(dir_path) if f.lower().endswith(valid_extensions) and os.path.isfile(os.path.join(dir_path, f))]
            self.image_files = sorted(files)
        except Exception as e:
            self.image_files = []
            if not silent:
                messagebox.showerror("Error", f"Could not read directory:\n{str(e)}")
        if self.image_files:
            self.current_preview_index = 0
            self.load_preview_image()
            if not silent:
                self.status_label.config(text=f"Found {len(self.image_files)} shawl images in input directory.")
        else:
            self.current_preview_index = 0
            self.preview_image_raw = None
            self.preview_label.config(text="No matching images found in input folder.\nSupports: JPG, JPEG, PNG, BMP, WEBP")
            self.preview_counter_label.config(text="0 of 0 images")
            self.image_info_label.config(text="")
        self.update_nav_buttons()

    # ─────────────────────────────────────────────────────────────────────────
    # Preview System
    # ─────────────────────────────────────────────────────────────────────────
    def load_preview_image(self):
        if not self.image_files:
            return
        filename = self.image_files[self.current_preview_index]
        img_path = os.path.join(self.input_dir.get(), filename)
        try:
            self.preview_image_raw = Image.open(img_path)
            self.update_preview()
        except Exception as e:
            self.preview_label.config(text=f"Error loading preview for:\n{filename}\n{str(e)}")
            self.image_info_label.config(text="")

    def update_preview(self, *args):
        if self.preview_image_raw is None or not self.image_files:
            return
        filename = self.image_files[self.current_preview_index]
        filename_without_ext = os.path.splitext(filename)[0]
        text_lines = self._build_label_lines(filename_without_ext)
        full_text = "\n".join(text_lines)
        font_size = self.font_size_var.get()
        text_color = self.text_color_var.get()
        bg_color = self.bg_color_var.get()
        bg_opacity = self.bg_opacity_var.get()
        position = self.position_var.get()
        padding = self.padding_var.get()
        margin = self.margin_var.get()
        font_file = FONTS.get(self.font_name_var.get(), "arialbd.ttf")
        img_copy = self.preview_image_raw.copy()
        labeled_img = self.apply_label_to_image(img_copy, full_text, font_file, font_size, text_color, bg_color, bg_opacity, position, padding, margin)
        container_w = self.preview_canvas_frame.winfo_width()
        container_h = self.preview_canvas_frame.winfo_height()
        if container_w <= 1 or container_h <= 1:
            container_w = 600
            container_h = 450
        img_w, img_h = labeled_img.size
        scale = min(container_w / img_w, container_h / img_h)
        new_w = max(10, int(img_w * scale))
        new_h = max(10, int(img_h * scale))
        resized_img = labeled_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(resized_img)
        self.preview_label.config(image=photo, text="")
        self.preview_label.image = photo
        self.preview_counter_label.config(text=f"Image {self.current_preview_index + 1} of {len(self.image_files)}")
        self.image_info_label.config(text=f"{filename} ({img_w}x{img_h})")

    def _build_label_lines(self, art_no_str):
        template = self.template_var.get()
        first_line = template.replace("{filename}", art_no_str)
        lines = [first_line]
        # Extract numeric part of art number, ignoring any alphabetic characters
        numeric_art_no = ''.join(filter(str.isdigit, art_no_str))
        # Use numeric part for lookup; fallback to original if no digits
        lookup_key = numeric_art_no if numeric_art_no else art_no_str
        row_data = self.xlsx_data.get(lookup_key)
        if row_data is None:
            try:
                art_int = int(lookup_key)
                row_data = self.xlsx_data.get(str(art_int))
            except (ValueError, TypeError):
                pass
        for field in self.extra_fields:
            prefix = field["text_var"].get().strip()
            col_name = field["col_var"].get()
            if col_name and col_name != "— load xlsx first —" and row_data is not None:
                value = row_data.get(col_name, "")
                value_str = "" if value is None else str(value).strip()
                line = f"{prefix} {value_str}" if prefix else value_str
                lines.append(line)
            elif prefix:
                lines.append(prefix)
        return lines

    def schedule_preview_update(self, *args):
        if self.preview_debounce_id:
            self.root.after_cancel(self.preview_debounce_id)
        self.preview_debounce_id = self.root.after(150, self.update_preview)

    # ─────────────────────────────────────────────────────────────────────────
    # Core Image Labeling
    # ─────────────────────────────────────────────────────────────────────────
    def apply_label_to_image(self, img, text, font_file, font_size, text_color, bg_color, bg_opacity, position, padding, margin):
        img_rgba = img.convert("RGBA")
        overlay = Image.new("RGBA", img_rgba.size, (0, 0, 0, 0))
        draw_overlay = ImageDraw.Draw(overlay)
        font_path = font_file
        if not os.path.exists(font_path):
            win_font_path = os.path.join("C:\\Windows\\Fonts", font_file)
            if os.path.exists(win_font_path):
                font_path = win_font_path
        try:
            font = ImageFont.truetype(font_path, font_size)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()
        lines = text.split("\n") if text else [""]
        line_bboxes = []
        for line in lines:
            try:
                bb = draw_overlay.textbbox((0, 0), line, font=font)
            except AttributeError:
                w, h = (len(line) * font_size // 2, font_size)
                bb = (0, 0, w, h)
            line_bboxes.append(bb)
        max_text_w = max((bb[2] - bb[0]) for bb in line_bboxes) if line_bboxes else 0
        try:
            sample_bb = draw_overlay.textbbox((0, 0), "Ag", font=font)
            line_h = sample_bb[3] - sample_bb[1]
        except Exception:
            line_h = font_size
        line_spacing = int(line_h * 1.35)
        total_text_h = line_h + (len(lines) - 1) * line_spacing if lines else line_h
        width, height = img_rgba.size
        box_w = max_text_w + 2 * padding
        box_h = total_text_h + 2 * padding
        if position == "Top-Left":
            bg_x = margin
            bg_y = margin
        elif position == "Top-Right":
            bg_x = width - margin - box_w
            bg_y = margin
        elif position == "Bottom-Left":
            bg_x = margin
            bg_y = height - margin - box_h
        elif position == "Bottom-Right":
            bg_x = width - margin - box_w
            bg_y = height - margin - box_h
        elif position == "Center":
            bg_x = (width - box_w) // 2
            bg_y = (height - box_h) // 2
        elif position == "Top-Center":
            bg_x = (width - box_w) // 2
            bg_y = margin
        elif position == "Bottom-Center":
            bg_x = (width - box_w) // 2
            bg_y = height - margin - box_h
        else:
            bg_x = margin
            bg_y = margin
        bg_rgb = hex_to_rgb(bg_color)
        bg_alpha = int(bg_opacity * 255)
        draw_overlay.rectangle([bg_x, bg_y, bg_x + box_w, bg_y + box_h], fill=(bg_rgb[0], bg_rgb[1], bg_rgb[2], bg_alpha))
        img_combined = Image.alpha_composite(img_rgba, overlay)
        draw_final = ImageDraw.Draw(img_combined)
        text_rgb = hex_to_rgb(text_color)
        for i, (line, bb) in enumerate(zip(lines, line_bboxes)):
            line_w = bb[2] - bb[0]
            offset_x = bb[0]
            offset_y = bb[1]
            x = bg_x + padding + (max_text_w - line_w) // 2 - offset_x
            y = bg_y + padding + i * line_spacing - offset_y
            draw_final.text((x, y), line, font=font, fill=(text_rgb[0], text_rgb[1], text_rgb[2], 255))
        return img_combined.convert("RGB")

    # ─────────────────────────────────────────────────────────────────────────
    # Navigation
    # ─────────────────────────────────────────────────────────────────────────
    def show_prev_image(self):
        if self.image_files and self.current_preview_index > 0:
            self.current_preview_index -= 1
            self.load_preview_image()
            self.update_nav_buttons()

    def show_next_image(self):
        if self.image_files and self.current_preview_index < len(self.image_files) - 1:
            self.current_preview_index += 1
            self.load_preview_image()
            self.update_nav_buttons()

    def update_nav_buttons(self):
        if not self.image_files:
            self.btn_prev.config(state="disabled")
            self.btn_next.config(state="disabled")
        else:
            self.btn_prev.config(state="normal" if self.current_preview_index > 0 else "disabled")
            self.btn_next.config(state="normal" if self.current_preview_index < len(self.image_files) - 1 else "disabled")

    # ─────────────────────────────────────────────────────────────────────────
    # Color Pickers
    # ─────────────────────────────────────────────────────────────────────────
    def choose_font_color(self):
        color_code = colorchooser.askcolor(initialcolor=self.text_color_var.get())
        if color_code[1]:
            self.text_color_var.set(color_code[1])
            self.font_color_preview.config(bg=color_code[1])
            self.schedule_preview_update()

    def choose_bg_color(self):
        color_code = colorchooser.askcolor(initialcolor=self.bg_color_var.get())
        if color_code[1]:
            self.bg_color_var.set(color_code[1])
            self.bg_color_preview.config(bg=color_code[1])
            self.schedule_preview_update()

    # ─────────────────────────────────────────────────────────────────────────
    # Batch Processing
    # ─────────────────────────────────────────────────────────────────────────
    def start_batch_processing(self):
        if self.processing:
            return
        input_dir = self.input_dir.get()
        output_dir = self.output_dir.get()
        if not input_dir or not os.path.exists(input_dir):
            messagebox.showerror("Error", "Please select a valid Input Folder.")
            return
        if not output_dir:
            messagebox.showerror("Error", "Please select an Output Folder.")
            return
        if not self.image_files:
            messagebox.showerror("Error", "No image files found to process.")
            return
        try:
            os.makedirs(output_dir, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not create output folder:\n{str(e)}")
            return
        self.processing = True
        self.btn_process.config(state="disabled", text="PROCESSING SHAWLS...")
        self.status_label.config(text="Processing images... Please wait.")
        fields_snapshot = [{"prefix": f["text_var"].get().strip(), "col": f["col_var"].get()} for f in self.extra_fields]
        thread = threading.Thread(target=self.run_batch_processing, args=(fields_snapshot,))
        thread.daemon = True
        thread.start()

    def run_batch_processing(self, fields_snapshot):
        input_dir = self.input_dir.get()
        output_dir = self.output_dir.get()
        font_size = self.font_size_var.get()
        text_color = self.text_color_var.get()
        bg_color = self.bg_color_var.get()
        bg_opacity = self.bg_opacity_var.get()
        position = self.position_var.get()
        template = self.template_var.get()
        padding = self.padding_var.get()
        margin = self.margin_var.get()
        font_file = FONTS.get(self.font_name_var.get(), "arialbd.ttf")
        xlsx_data_snap = dict(self.xlsx_data)
        total_files = len(self.image_files)
        processed_count = 0
        error_count = 0
        error_list = []
        self.progress["maximum"] = total_files
        self.progress["value"] = 0
        for index, filename in enumerate(self.image_files):
            try:
                self.root.after(0, self.update_processing_status, f"Processing ({index+1}/{total_files}): {filename}", index)
                img_path = os.path.join(input_dir, filename)
                img = Image.open(img_path)
                filename_without_ext = os.path.splitext(filename)[0]
                first_line = template.replace("{filename}", filename_without_ext)
                lines = [first_line]
                # Strip alphabetic characters for Excel lookup (e.g. "423A" -> "423")
                numeric_art_no = ''.join(filter(str.isdigit, filename_without_ext))
                lookup_key = numeric_art_no if numeric_art_no else filename_without_ext
                row_data = xlsx_data_snap.get(lookup_key)
                if row_data is None:
                    try:
                        art_int = int(lookup_key)
                        row_data = xlsx_data_snap.get(str(art_int))
                    except (ValueError, TypeError):
                        pass
                for field in fields_snapshot:
                    prefix = field["prefix"]
                    col_name = field["col"]
                    if col_name and col_name != "— load xlsx first —" and row_data is not None:
                        value = row_data.get(col_name, "")
                        value_str = "" if value is None else str(value).strip()
                        lines.append(f"{prefix} {value_str}" if prefix else value_str)
                    elif prefix:
                        lines.append(prefix)
                full_text = "\n".join(lines)
                labeled_img = self.apply_label_to_image(img, full_text, font_file, font_size, text_color, bg_color, bg_opacity, position, padding, margin)
                out_path = os.path.join(output_dir, filename)
                ext = os.path.splitext(filename)[1].lower()
                save_format = "JPEG" if ext in ['.jpg', '.jpeg'] else ext.replace('.', '').upper()
                if save_format == "JPG":
                    save_format = "JPEG"
                labeled_img.save(out_path, format=save_format, quality=95)
                processed_count += 1
            except Exception as e:
                error_count += 1
                error_list.append(f"{filename}: {str(e)}")
        self.root.after(0, self.on_processing_complete, processed_count, error_count, error_list)

    def update_processing_status(self, text, value):
        self.status_label.config(text=text)
        self.progress["value"] = value

    def on_processing_complete(self, processed, errors, error_list):
        self.processing = False
        self.btn_process.config(state="normal", text="PROCESS ALL SHAWL IMAGES")
        self.progress["value"] = len(self.image_files)
        self.status_label.config(text=f"Completed! Labeled {processed} shawls. Errors: {errors}")
        if errors > 0:
            error_details = "\n".join(error_list[:5])
            if len(error_list) > 5:
                error_details += "\n...and more."
            messagebox.showwarning(
                "Processing Complete with Errors",
                f"Successfully processed {processed} images.\nFailed to process {errors} images.\n\nErrors:\n{error_details}"
            )
        else:
            messagebox.showinfo(
                "Success!",
                f"Successfully processed all {processed} shawl images!\nLabeled output saved to folder:\n{self.output_dir.get()}"
            )
        try:
            os.startfile(self.output_dir.get())
        except Exception:
            pass


if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    root = tk.Tk()
    app = LabelerApp(root)
    def on_window_resize(event):
        if event.widget == app.preview_canvas_frame:
            app.schedule_preview_update()
    app.preview_canvas_frame.bind("<Configure>", on_window_resize)
    root.mainloop()
